"""
report_builder.py
Tạo file Excel (.xlsx) báo cáo điểm danh đẹp cho giảng viên.
"""
import os
import uuid
from datetime import datetime, date, time
from decimal import Decimal
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ─── Thư mục lưu file ─────────────────────────────────────────────────────────
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "static", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


# ─── Bảng màu ─────────────────────────────────────────────────────────────────
C_DARK_NAVY  = "1A1A2E"   # header nền
C_WHITE      = "FFFFFF"
C_LIGHT_GRAY = "F4F6F9"
C_MID_GRAY   = "E2E8F0"
C_BORDER     = "CBD5E1"

C_GREEN_BG   = "DCFCE7"; C_GREEN_FG  = "15803D"
C_AMBER_BG   = "FEF3C7"; C_AMBER_FG  = "B45309"
C_RED_BG     = "FEE2E2"; C_RED_FG    = "B91C1C"
C_BLUE_BG    = "DBEAFE"; C_BLUE_FG   = "1D4ED8"
C_PURPLE_BG  = "EDE9FE"; C_PURPLE_FG = "6D28D9"

C_TH_BG      = "334155"   # table header
C_TH_FG      = "FFFFFF"


# ─── Helpers style ────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="1A1A2E", italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

def _border_thin() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium() -> Border:
    m = Side(style="medium", color=C_DARK_NAVY)
    s = Side(style="thin", color=C_BORDER)
    return Border(left=m, right=m, top=m, bottom=m)

def _apply(cell, fill=None, font=None, alignment=None, border=None, number_format=None):
    if fill:         cell.fill = fill
    if font:         cell.font = font
    if alignment:    cell.alignment = alignment
    if border:       cell.border = border
    if number_format: cell.number_format = number_format


# ─── Tiện ích ─────────────────────────────────────────────────────────────────
def _fmt_val(v):
    """Chuẩn hoá giá trị cho cell."""
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y") if isinstance(v, date) else v.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, Decimal):
        return float(v)
    return v


def _write_header_block(ws, row: int, title: str, subtitle: str = ""):
    """Dòng tiêu đề lớn màu navy."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value=title)
    _apply(c, fill=_fill(C_DARK_NAVY), font=_font(bold=True, size=14, color=C_WHITE), alignment=_center())
    ws.row_dimensions[row].height = 34

    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=10)
        cs = ws.cell(row=row+1, column=1, value=subtitle)
        _apply(cs, fill=_fill("2D2D4E"), font=_font(size=10, color="A0AEC0"), alignment=_center())
        ws.row_dimensions[row+1].height = 22
        return row + 2
    return row + 1


def _write_metric_row(ws, row: int, metrics: list):
    """
    metrics = [(label, value, bg, fg), ...]  tối đa 5 metric
    Mỗi metric chiếm 2 cột (label trên, value dưới).
    """
    col = 1
    for label, value, bg, fg in metrics:
        # Label
        lc = ws.cell(row=row, column=col, value=label)
        _apply(lc, fill=_fill(bg), font=_font(size=9, color=fg), alignment=_center())
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        # Value
        vc = ws.cell(row=row+1, column=col, value=value)
        _apply(vc, fill=_fill(bg), font=_font(bold=True, size=20, color=fg), alignment=_center())
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        ws.row_dimensions[row].height   = 18
        ws.row_dimensions[row+1].height = 36
        col += 2
    return row + 2


def _write_section_title(ws, row: int, col: int, text: str, col_end: int = 10):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col, value=text)
    _apply(c, fill=_fill(C_MID_GRAY), font=_font(bold=True, size=11, color=C_DARK_NAVY), alignment=_left())
    ws.row_dimensions[row].height = 26
    return row + 1


def _write_table_header(ws, row: int, cols: list):
    """cols = [(header_text, col_width), ...]"""
    for i, (header, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=header)
        _apply(c, fill=_fill(C_TH_BG), font=_font(bold=True, size=10, color=C_TH_FG), alignment=_center(), border=_border_thin())
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 26
    return row + 1


def _status_style(status: str):
    if status == "Đúng giờ":
        return "Đúng giờ", C_GREEN_BG, C_GREEN_FG
    if status == "Đi muộn":
        return "Đi muộn", C_AMBER_BG, C_AMBER_FG
    return "Vắng", C_RED_BG, C_RED_FG


# ══════════════════════════════════════════════════════════════════════════════
# BÁO CÁO 1 LỚP
# ══════════════════════════════════════════════════════════════════════════════
def build_class_report_xlsx(session_info: dict, detail_rows: List[dict]) -> str:
    """
    Tạo Excel báo cáo điểm danh 1 lớp.
    Trả về đường dẫn file đã lưu.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo điểm danh"
    ws.sheet_view.showGridLines = False

    # ── Thông tin buổi học ────────────────────────────────────────────────────
    si        = session_info or {}
    class_code = si.get("class_code", "N/A")
    class_name = si.get("class_name", "")
    date_str   = si.get("date", "")
    if isinstance(date_str, date):
        date_str = date_str.strftime("%d/%m/%Y")
    st = si.get("start_time", "")
    if isinstance(st, time):
        st = st.strftime("%H:%M")
    et = si.get("end_time", "")
    if isinstance(et, time):
        et = et.strftime("%H:%M")
    threshold  = si.get("late_threshold_minutes", "N/A")
    generated  = datetime.now().strftime("%H:%M %d/%m/%Y")

    # ── Tính thống kê ─────────────────────────────────────────────────────────
    total   = len(detail_rows)
    on_time = sum(1 for r in detail_rows if r.get("Trạng thái") == "Đúng giờ")
    late    = sum(1 for r in detail_rows if r.get("Trạng thái") == "Đi muộn")
    absent  = sum(1 for r in detail_rows if r.get("Trạng thái") == "Vắng")
    rate    = round((on_time + late) * 100 / total, 1) if total else 0

    row = 1

    # ── Block tiêu đề ─────────────────────────────────────────────────────────
    subtitle = (
        f"Môn: {class_name}   |   Ngày: {date_str}   |   "
        f"Giờ: {st} – {et}   |   Trễ sau: {threshold} phút   |   Tạo lúc: {generated}"
    )
    row = _write_header_block(ws, row, f"BÁO CÁO ĐIỂM DANH  —  {class_code}", subtitle)
    row += 1  # khoảng cách

    # ── Metric cards ──────────────────────────────────────────────────────────
    row = _write_metric_row(ws, row, [
        ("TỔNG SINH VIÊN", total,   C_BLUE_BG,   C_BLUE_FG),
        ("ĐÚNG GIỜ",       on_time, C_GREEN_BG,  C_GREEN_FG),
        ("ĐI MUỘN",        late,    C_AMBER_BG,  C_AMBER_FG),
        ("VẮNG",           absent,  C_RED_BG,    C_RED_FG),
        ("TỈ LỆ CÓ MẶT",  f"{rate}%", C_PURPLE_BG, C_PURPLE_FG),
    ])
    row += 1

    # ── Bảng chi tiết sinh viên ───────────────────────────────────────────────
    row = _write_section_title(ws, row, 1, "  CHI TIẾT TỪNG SINH VIÊN")
    row = _write_table_header(ws, row, [
        ("STT", 5), ("Mã SV", 12), ("Họ và tên", 28),
        ("Trạng thái", 14), ("Giờ quét thẻ", 14),
    ])

    for i, r in enumerate(detail_rows, 1):
        status_label, bg, fg = _status_style(r.get("Trạng thái", "Vắng"))
        scan = r.get("Giờ quét") or "—"
        data_row = [i, r.get("Mã SV", ""), r.get("Họ tên", ""), status_label, scan]
        for col_idx, val in enumerate(data_row, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            row_bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
            if col_idx == 4:   # cột trạng thái → màu đặc biệt
                _apply(c, fill=_fill(bg), font=_font(bold=True, size=10, color=fg), alignment=_center(), border=_border_thin())
            elif col_idx == 1:
                _apply(c, fill=_fill(row_bg), font=_font(size=10, color="9CA3AF"), alignment=_center(), border=_border_thin())
            else:
                _apply(c, fill=_fill(row_bg), font=_font(size=10), alignment=_left() if col_idx == 3 else _center(), border=_border_thin())
        ws.row_dimensions[row].height = 22
        row += 1

    if not detail_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value="Chưa có dữ liệu sinh viên")
        _apply(c, font=_font(italic=True, color="9CA3AF"), alignment=_center())

    row += 1

    # ── Chú thích ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=f"* Báo cáo được tạo tự động lúc {generated}")
    _apply(c, font=_font(size=9, color="9CA3AF", italic=True), alignment=_left())

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A8"

    # ── Lưu file ──────────────────────────────────────────────────────────────
    filename = f"BaoCao_{class_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    return filepath, f"/static/reports/{filename}", filename


# ══════════════════════════════════════════════════════════════════════════════
# BÁO CÁO CẢ NGÀY (TẤT CẢ LỚP)
# ══════════════════════════════════════════════════════════════════════════════
def build_daily_report_xlsx(
    date_label: str,
    summary_rows: List[dict],
    late_rows: List[dict],
    absent_rows: List[dict],
) -> str:
    wb = Workbook()
    generated = datetime.now().strftime("%H:%M %d/%m/%Y")

    # ══ Sheet 1: Tổng hợp ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    ws1.sheet_view.showGridLines = False

    total_classes = len(summary_rows)
    total_sv      = sum(int(r.get("tong_sv") or 0)  for r in summary_rows)
    total_on_time = sum(int(r.get("on_time")  or 0) for r in summary_rows)
    total_late    = sum(int(r.get("late")     or 0) for r in summary_rows)
    total_absent  = sum(int(r.get("absent")   or 0) for r in summary_rows)
    overall_rate  = round((total_on_time + total_late) * 100 / total_sv, 1) if total_sv else 0

    row = 1
    row = _write_header_block(ws1, row,
        f"BÁO CÁO ĐIỂM DANH  —  {date_label.upper()}",
        f"Tổng hợp toàn hệ thống   |   Tạo lúc: {generated}"
    )
    row += 1

    row = _write_metric_row(ws1, row, [
        ("SỐ LỚP HỌC",      total_classes, C_BLUE_BG,   C_BLUE_FG),
        ("TỔNG LƯỢT SV",    total_sv,      C_BLUE_BG,   C_BLUE_FG),
        ("ĐÚNG GIỜ",        total_on_time, C_GREEN_BG,  C_GREEN_FG),
        ("ĐI MUỘN",         total_late,    C_AMBER_BG,  C_AMBER_FG),
        ("VẮNG",            total_absent,  C_RED_BG,    C_RED_FG),
    ])
    # Thêm ô tỉ lệ riêng bên dưới
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cr = ws1.cell(row=row, column=1, value=f"Tỉ lệ điểm danh trung bình toàn hệ thống:  {overall_rate}%")
    _apply(cr, fill=_fill(C_PURPLE_BG), font=_font(bold=True, size=12, color=C_PURPLE_FG), alignment=_center())
    ws1.row_dimensions[row].height = 28
    row += 2

    row = _write_section_title(ws1, row, 1, "  TỔNG HỢP TỪNG LỚP")
    cols_summary = [
        ("Mã lớp", 10), ("Tên môn học", 28), ("Ngày", 12), ("Thời gian", 16),
        ("Tổng SV", 10), ("Đúng giờ", 10), ("Đi muộn", 10),
        ("Vắng", 10), ("Tỉ lệ %", 10),
    ]
    row = _write_table_header(ws1, row, cols_summary)

    for i, r in enumerate(summary_rows, 1):
        rate_val = r.get("rate")
        try:
            rate_f = float(rate_val)
        except (TypeError, ValueError):
            rate_f = 0
        rate_str = f"{rate_f}%"

        if rate_f >= 90:
            rate_bg, rate_fg = C_GREEN_BG, C_GREEN_FG
        elif rate_f >= 75:
            rate_bg, rate_fg = C_AMBER_BG, C_AMBER_FG
        else:
            rate_bg, rate_fg = C_RED_BG, C_RED_FG

        row_bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
        # Combine start and end time
        start_time = r.get("start_time", "")
        end_time = r.get("end_time", "")
        time_range = f"{start_time} - {end_time}" if start_time and end_time else start_time
        vals = [
            r.get("class_code", ""), r.get("class_name", ""), r.get("date", ""),
            time_range, r.get("tong_sv", 0),
            r.get("on_time", 0), r.get("late", 0),
            r.get("absent", 0), rate_str,
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col_idx, value=val)
            if col_idx == 9:
                _apply(c, fill=_fill(rate_bg), font=_font(bold=True, size=10, color=rate_fg), alignment=_center(), border=_border_thin())
            elif col_idx == 6:
                _apply(c, fill=_fill(row_bg), font=_font(size=10, color=C_GREEN_FG), alignment=_center(), border=_border_thin())
            elif col_idx == 7:
                _apply(c, fill=_fill(row_bg), font=_font(size=10, color=C_AMBER_FG), alignment=_center(), border=_border_thin())
            elif col_idx == 8:
                _apply(c, fill=_fill(row_bg), font=_font(size=10, color=C_RED_FG), alignment=_center(), border=_border_thin())
            else:
                _apply(c, fill=_fill(row_bg), font=_font(size=10), alignment=_left() if col_idx in (2,3) else _center(), border=_border_thin())
        ws1.row_dimensions[row].height = 22
        row += 1

    if not summary_rows:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws1.cell(row=row, column=1, value="Không có lớp học nào trong ngày này")
        _apply(c, font=_font(italic=True, color="9CA3AF"), alignment=_center())

    ws1.freeze_panes = "A8"

    # ══ Sheet 2: Sinh viên đi muộn ════════════════════════════════════════════
    ws2 = wb.create_sheet("Đi muộn")
    ws2.sheet_view.showGridLines = False

    row = 1
    row = _write_header_block(ws2, row,
        f"DANH SÁCH ĐI MUỘN  —  {date_label.upper()}",
        f"Tổng: {len(late_rows)} sinh viên   |   Tạo lúc: {generated}"
    )
    row += 1

    if late_rows:
        row = _write_table_header(ws2, row, [
            ("STT", 5), ("Mã SV", 12), ("Họ và tên", 28),
            ("Mã lớp", 10), ("Tên môn học", 28), ("Ngày", 12), ("Giờ quét thẻ", 14),
        ])
        for i, r in enumerate(late_rows, 1):
            row_bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
            vals = [i, r.get("student_code",""), r.get("full_name",""),
                    r.get("class_code",""), r.get("class_name",""), r.get("date",""), r.get("scan_time","—")]
            for col_idx, val in enumerate(vals, 1):
                c = ws2.cell(row=row, column=col_idx, value=val)
                _apply(c, fill=_fill(row_bg), font=_font(size=10),
                       alignment=_left() if col_idx in (3,5) else _center(), border=_border_thin())
            ws2.row_dimensions[row].height = 22
            row += 1
    else:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws2.cell(row=row, column=1, value="Không có sinh viên đi muộn")
        _apply(c, font=_font(italic=True, color="9CA3AF"), alignment=_center())

    ws2.freeze_panes = "A5"

    # ══ Sheet 3: Sinh viên vắng ════════════════════════════════════════════════
    ws3 = wb.create_sheet("Vắng mặt")
    ws3.sheet_view.showGridLines = False

    row = 1
    row = _write_header_block(ws3, row,
        f"DANH SÁCH VẮNG MẶT  —  {date_label.upper()}",
        f"Tổng: {len(absent_rows)} sinh viên   |   Tạo lúc: {generated}"
    )
    row += 1

    if absent_rows:
        row = _write_table_header(ws3, row, [
            ("STT", 5), ("Mã SV", 12), ("Họ và tên", 28),
            ("Mã lớp", 10), ("Tên môn học", 28), ("Ngày", 12),
        ])
        for i, r in enumerate(absent_rows, 1):
            row_bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
            vals = [i, r.get("student_code",""), r.get("full_name",""),
                    r.get("class_code",""), r.get("class_name",""), r.get("date","")]
            for col_idx, val in enumerate(vals, 1):
                c = ws3.cell(row=row, column=col_idx, value=val)
                _apply(c, fill=_fill(row_bg), font=_font(size=10),
                       alignment=_left() if col_idx in (3,5) else _center(), border=_border_thin())
            ws3.row_dimensions[row].height = 22
            row += 1
    else:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws3.cell(row=row, column=1, value="Không có sinh viên vắng")
        _apply(c, font=_font(italic=True, color="9CA3AF"), alignment=_center())

    ws3.freeze_panes = "A5"

    # ── Lưu file ──────────────────────────────────────────────────────────────
    safe_label = date_label.replace(" ", "_").replace("/", "-")
    filename = f"BaoCao_NgayHoc_{safe_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    return filepath, f"/static/reports/{filename}", filename